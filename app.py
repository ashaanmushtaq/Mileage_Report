import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill, NamedStyle
from openpyxl.utils import get_column_letter
import io
import re
from datetime import datetime

# ------------------------------------------------------------------------------
# DEFAULT RURAL VEHICLE MASTER LIST
# ------------------------------------------------------------------------------
DEFAULT_RURAL_LIST = [
    # Trolly / Other Vehicles (17)
    "GAB-4046", "SA-6766", "STR-6637", "TT-06", "GTD-694", "GAS-1694",
    "BN-3932", "DGK-1763", "TT-11", "GAU-8135", "GAJ-19-47", "SAA-2946",
    "ST-663", "GAJ-3943", "TT-05", "OKA-2192", "SAA-7988",
    
    # R Series Rickshaws (21)
    "R-1", "R-2", "R-3", "R-4", "R-5", "R-6", "R-7", "R-8", "R-9", 
    "R-11", "R-12", "R-13", "R-14", "R-15", "R-16", "R-17", "R-23", 
    "R-32", "R-33", "R-35", "R-41",
    
    # RIC Series Rickshaws (9)
    "RIC-26", "RIC-28", "RIC-31", "RIC-34", "RIC-37", "RIC-39", 
    "RIC-42", "RIC-43", "RIC-44"
]

st.set_page_config(
    page_title="Mileage Processor | Dev by Ashaan",
    page_icon="🧾",
    layout="wide"
)

# ------------------------------------------------------------------------------
# AESTHETIC RECEIPT UI STYLING
# ------------------------------------------------------------------------------
st.markdown("""
    <style>
    .receipt-card {
        background-color: #FFFFFF;
        border: 2px dashed #1F4E79;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
        margin-bottom: 20px;
    }
    .receipt-header {
        font-family: 'Courier New', Courier, monospace;
        font-size: 26px;
        font-weight: bold;
        color: #1F4E79;
        text-align: center;
        letter-spacing: 2px;
        margin-bottom: 5px;
    }
    .receipt-sub {
        font-family: 'Calibri', sans-serif;
        text-align: center;
        color: #555;
        font-size: 14px;
        margin-bottom: 15px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="receipt-card">
    <div class="receipt-header">🧾 MILEAGE EXECUTIVE AUDIT DASHBOARD</div>
    <div class="receipt-sub">Developed by: <b>Muhammad Ashaan</b> | Bellanix Tech</div>
</div>
""", unsafe_allow_html=True)

# ------------------------------------------------------------------------------
# HELPER FUNCTIONS
# ------------------------------------------------------------------------------
def parse_period_to_hours(val):
    if pd.isna(val) or val is None:
        return -1.0
    val_str = str(val).replace('\xa0', '').strip()
    if not val_str:
        return -1.0
    
    if ':' in val_str:
        parts = val_str.split(':')
        try:
            hrs = float(parts[0])
            mins = float(parts[1]) if len(parts) > 1 else 0.0
            secs = float(parts[2]) if len(parts) > 2 else 0.0
            return hrs + (mins / 60.0) + (secs / 3600.0)
        except:
            return -1.0
    try:
        f = float(val_str)
        return f * 24.0 if f <= 1.0 else f
    except:
        return -1.0

def normalize_reg(val):
    if pd.isna(val):
        return ""
    return str(val).replace('\xa0', '').replace('-', '').replace(' ', '').strip().upper()

# STRICT MATCHING FUNCTION: Sirf wahi rural me jayega jo master list me bilkul match hoga
def is_rural_vehicle(reg_str, normalized_rural_set):
    norm_clean = normalize_reg(reg_str)
    return norm_clean in normalized_rural_set

def extract_file_date(df_raw):
    for r in range(min(5, len(df_raw))):
        row_str = " ".join(df_raw.iloc[r].dropna().astype(str).tolist())
        if '[' in row_str and ']' in row_str:
            p1, p2 = row_str.find('['), row_str.find(']')
            extracted = row_str[p1+1:p2].split('-')[0].strip()
            if len(extracted) >= 6:
                return extracted
        match = re.search(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b', row_str)
        if match:
            return match.group(0)
    return datetime.now().strftime("%m/%d/%Y")

def process_mileage_df(df_raw):
    header_row_idx, col_reg_idx, col_period_idx = -1, -1, -1
    for r_idx in range(min(15, len(df_raw))):
        row_vals = df_raw.iloc[r_idx].astype(str).str.replace('\xa0', '', regex=False).str.strip().str.upper()
        c_reg, c_per = -1, -1
        for c_idx, val in enumerate(row_vals):
            if val in ['REG#', 'REGISTRATION', 'VEHICLE NO', 'REGISTRATION NO', 'REG NO']:
                c_reg = c_idx
            if val in ['PERIOD', 'DURATION', 'HOURS']:
                c_per = c_idx
        if c_reg != -1 and c_per != -1:
            header_row_idx, col_reg_idx, col_period_idx = r_idx, c_reg, c_per
            break
    return header_row_idx, col_reg_idx, col_period_idx

# ------------------------------------------------------------------------------
# ENHANCED EXCEL REPORT GENERATOR
# ------------------------------------------------------------------------------
def generate_professional_excel(headers, urban_rows, rural_rows, selected_city, meta_date, 
                                 total_valid_vehicles, repeated_vehicles, duplicate_regs, 
                                 has_sno, c_reg_idx, prev_vehicles):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Report"
    
    ws.sheet_view.showGridLines = False
    
    enhanced_headers = headers.copy()
    enhanced_headers.append("STATUS")
    num_cols = len(enhanced_headers)
    half_cols = max(2, num_cols // 2)
    
    COLORS = {
        'primary_dark': '1C2E4A',
        'primary_medium': '2C4A6E',
        'primary_light': '3B6B9E',
        'header_bg': 'F0F4F8',
        'urban_bg': 'E8F0FE',
        'rural_bg': 'E8F5E9',
        'border_color': 'D0D7E6',
        'text_dark': '1A2332',
        'text_light': 'FFFFFF',
        'warning_bg': 'FFF3CD',
        'repeat_bg': 'FFE5CC',
        'new_bg': 'D4EDDA',
        'alt_row': 'F8FAFC'
    }
    
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(name='Calibri', size=18, bold=True, color=COLORS['text_light'])
    title_style.fill = PatternFill(start_color=COLORS['primary_dark'], end_color=COLORS['primary_dark'], fill_type="solid")
    title_style.alignment = Alignment(horizontal='center', vertical='center')
    
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name='Calibri', size=11, bold=True, color=COLORS['text_light'])
    header_style.fill = PatternFill(start_color=COLORS['primary_medium'], end_color=COLORS['primary_medium'], fill_type="solid")
    header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_style.border = Border(bottom=Side(style='medium', color=COLORS['primary_dark']))
    
    wb.add_named_style(title_style)
    wb.add_named_style(header_style)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border_color']),
        right=Side(style='thin', color=COLORS['border_color']),
        top=Side(style='thin', color=COLORS['border_color']),
        bottom=Side(style='thin', color=COLORS['border_color'])
    )
    
    # 1. Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{selected_city} VEHICLE MILEAGE EXECUTIVE REPORT")
    title_cell.style = 'title_style'
    ws.row_dimensions[1].height = 45
    
    # 2. Subtitle Lines
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sep_cell = ws.cell(row=2, column=1)
    sep_cell.fill = PatternFill(start_color=COLORS['primary_dark'], end_color=COLORS['primary_dark'], fill_type="solid")
    ws.row_dimensions[2].height = 3
    
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=half_cols)
    r3_left = ws.cell(row=3, column=1, value=f"📅 REPORT DATE:  {meta_date}")
    r3_left.font = Font(name='Calibri', size=11, bold=True, color=COLORS['primary_dark'])
    r3_left.alignment = center_align
    r3_left.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    
    ws.merge_cells(start_row=3, start_column=half_cols+1, end_row=3, end_column=num_cols)
    r3_right = ws.cell(row=3, column=half_cols+1, value=f"🚗 TOTAL FLEET (20-24h):  {total_valid_vehicles}")
    r3_right.font = Font(name='Calibri', size=11, bold=True, color=COLORS['primary_dark'])
    r3_right.alignment = center_align
    r3_right.fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    ws.row_dimensions[3].height = 28
    
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=half_cols)
    r4_left = ws.cell(row=4, column=1, value="💻 DEVELOPED BY:  Muhammad Ashaan")
    r4_left.font = Font(name='Calibri', size=10, bold=True, color=COLORS['primary_medium'])
    r4_left.alignment = center_align
    
    ws.merge_cells(start_row=4, start_column=half_cols+1, end_row=4, end_column=num_cols)
    r4_right = ws.cell(row=4, column=half_cols+1, value=f"🏙️ URBAN FLEET:  {len(urban_rows)}")
    r4_right.font = Font(name='Calibri', size=10, bold=True, color='2C4A6E')
    r4_right.alignment = center_align
    ws.row_dimensions[4].height = 24
    
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=half_cols)
    r5_left = ws.cell(row=5, column=1, value=f"⏱️ FILTERED HOURS:  20 – 24 Hours")
    r5_left.font = Font(name='Calibri', size=10, color=COLORS['primary_dark'])
    r5_left.alignment = center_align
    
    ws.merge_cells(start_row=5, start_column=half_cols+1, end_row=5, end_column=num_cols)
    r5_right = ws.cell(row=5, column=half_cols+1, value=f"🌾 RURAL FLEET:  {len(rural_rows)}")
    r5_right.font = Font(name='Calibri', size=10, bold=True, color='1E7E34')
    r5_right.alignment = center_align
    ws.row_dimensions[5].height = 24
    
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=num_cols)
    sep2_cell = ws.cell(row=6, column=1)
    sep2_cell.fill = PatternFill(start_color=COLORS['primary_light'], end_color=COLORS['primary_light'], fill_type="solid")
    ws.row_dimensions[6].height = 2
    
    # 3. Table Headers
    header_row = 8
    ws.row_dimensions[header_row].height = 32
    for c_idx, h_text in enumerate(enhanced_headers, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=h_text)
        cell.style = 'header_style'
    
    # 4. Rows
    curr_row = 9
    sno_tracker = 1
    actual_reg_col = c_reg_idx + (0 if has_sno else 1)
    status_col = len(enhanced_headers)
    
    alt_fill = PatternFill(start_color=COLORS['alt_row'], end_color=COLORS['alt_row'], fill_type="solid")
    urban_bg_fill = PatternFill(start_color=COLORS['urban_bg'], end_color=COLORS['urban_bg'], fill_type="solid")
    rural_bg_fill = PatternFill(start_color=COLORS['rural_bg'], end_color=COLORS['rural_bg'], fill_type="solid")
    warning_fill = PatternFill(start_color=COLORS['warning_bg'], end_color=COLORS['warning_bg'], fill_type="solid")
    repeat_fill = PatternFill(start_color=COLORS['repeat_bg'], end_color=COLORS['repeat_bg'], fill_type="solid")
    new_fill = PatternFill(start_color=COLORS['new_bg'], end_color=COLORS['new_bg'], fill_type="solid")
    
    # Urban Section
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=num_cols)
    u_sec = ws.cell(row=curr_row, column=1, value=f"  🏙️  URBAN FLEET VEHICLES  ({len(urban_rows)})")
    u_sec.font = Font(name='Calibri', size=13, bold=True, color=COLORS['text_light'])
    u_sec.fill = PatternFill(start_color=COLORS['primary_medium'], end_color=COLORS['primary_medium'], fill_type="solid")
    u_sec.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[curr_row].height = 30
    curr_row += 1
    
    for idx, r_data in enumerate(urban_rows):
        reg_val = str(r_data[actual_reg_col]).replace('\xa0', '').strip() if pd.notna(r_data[actual_reg_col]) else ''
        reg_clean = reg_val.upper()
        reg_norm = normalize_reg(reg_clean)
        is_alt = (idx % 2 == 1)
        
        if reg_norm in prev_vehicles:
            status = "🔁 REPEATED"
            status_fill = repeat_fill
            status_font_color = 'B7410E'
        elif reg_clean in duplicate_regs:
            status = "⚠️ DUPLICATE"
            status_fill = warning_fill
            status_font_color = '9C5700'
        else:
            status = "✅ NEW"
            status_fill = new_fill
            status_font_color = '155724'
        
        r_data[0] = sno_tracker
        sno_tracker += 1
        row_with_status = r_data + [status]
        
        ws.row_dimensions[curr_row].height = 22
        for c_idx, val in enumerate(row_with_status, 1):
            val_clean = str(val).replace('\xa0', '').strip() if pd.notna(val) else ''
            cell = ws.cell(row=curr_row, column=c_idx, value=val_clean if c_idx > 1 else row_with_status[0])
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            
            cell.fill = alt_fill if is_alt else urban_bg_fill
            
            if c_idx == (actual_reg_col + 1):
                if reg_norm in prev_vehicles or reg_clean in duplicate_regs:
                    cell.font = Font(name='Calibri', size=10, bold=True, color=status_font_color)
                    cell.fill = status_fill
            
            if c_idx == status_col:
                cell.font = Font(name='Calibri', size=10, bold=True, color=status_font_color)
                cell.fill = status_fill
        curr_row += 1
    
    curr_row += 1
    
    # Rural Section
    ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=num_cols)
    r_sec = ws.cell(row=curr_row, column=1, value=f"  🌾  RURAL FLEET VEHICLES  ({len(rural_rows)})")
    r_sec.font = Font(name='Calibri', size=13, bold=True, color=COLORS['text_light'])
    r_sec.fill = PatternFill(start_color='1E7E34', end_color='1E7E34', fill_type="solid")
    r_sec.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[curr_row].height = 30
    curr_row += 1
    
    for idx, r_data in enumerate(rural_rows):
        reg_val = str(r_data[actual_reg_col]).replace('\xa0', '').strip() if pd.notna(r_data[actual_reg_col]) else ''
        reg_clean = reg_val.upper()
        reg_norm = normalize_reg(reg_clean)
        is_alt = (idx % 2 == 1)
        
        if reg_norm in prev_vehicles:
            status = "🔁 REPEATED"
            status_fill = repeat_fill
            status_font_color = 'B7410E'
        elif reg_clean in duplicate_regs:
            status = "⚠️ DUPLICATE"
            status_fill = warning_fill
            status_font_color = '9C5700'
        else:
            status = "✅ NEW"
            status_fill = new_fill
            status_font_color = '155724'
        
        r_data[0] = sno_tracker
        sno_tracker += 1
        row_with_status = r_data + [status]
        
        ws.row_dimensions[curr_row].height = 22
        for c_idx, val in enumerate(row_with_status, 1):
            val_clean = str(val).replace('\xa0', '').strip() if pd.notna(val) else ''
            cell = ws.cell(row=curr_row, column=c_idx, value=val_clean if c_idx > 1 else row_with_status[0])
            cell.alignment = center_align
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            
            cell.fill = alt_fill if is_alt else rural_bg_fill
            
            if c_idx == (actual_reg_col + 1):
                if reg_norm in prev_vehicles or reg_clean in duplicate_regs:
                    cell.font = Font(name='Calibri', size=10, bold=True, color=status_font_color)
                    cell.fill = status_fill
            
            if c_idx == status_col:
                cell.font = Font(name='Calibri', size=10, bold=True, color=status_font_color)
                cell.fill = status_fill
        curr_row += 1
    
    # Footer
    footer_row = curr_row + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=num_cols)
    footer_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    footer_cell = ws.cell(row=footer_row, column=1, value=footer_text)
    footer_cell.font = Font(name='Calibri', size=9, italic=True, color='6B7A8F')
    footer_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[footer_row].height = 20
    
    # Legend
    legend_row = footer_row + 2
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=num_cols)
    legend_cell = ws.cell(row=legend_row, column=1, value="📌 STATUS LEGEND:")
    legend_cell.font = Font(name='Calibri', size=11, bold=True, color=COLORS['primary_dark'])
    legend_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    legend_items = [
        ("🔁 REPEATED", "Vehicle appeared in previous day's 20-24h report", COLORS['repeat_bg'], 'B7410E'),
        ("✅ NEW", "New vehicle in today's 20-24h report", COLORS['new_bg'], '155724'),
        ("⚠️ DUPLICATE", "Same vehicle appears multiple times today", COLORS['warning_bg'], '9C5700'),
        ("🔵 Light Blue", "Urban fleet entry", COLORS['urban_bg'], '2C4A6E'),
        ("🟢 Light Green", "Rural fleet entry", COLORS['rural_bg'], '1E7E34')
    ]
    
    for i, (label, desc, color, text_color) in enumerate(legend_items):
        row = legend_row + 1 + i
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        
        legend_cell = ws.cell(row=row, column=1, value=f"  {label}:  {desc}")
        legend_cell.font = Font(name='Calibri', size=9, color=text_color)
        legend_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        if i < 5:
            color_cell = ws.cell(row=row, column=1)
            color_cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''), fill_type="solid")
    
    # Column Sizing & Locking
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        if col_letter == 'A':
            ws.column_dimensions['A'].width = 10
            continue
            
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 6, 14), 35)
    
    status_col_letter = get_column_letter(status_col)
    ws.column_dimensions[status_col_letter].width = 18
    
    ws.freeze_panes = 'A9'
    ws.protection.sheet = True
    ws.protection.password = 'Ashaan2026'
    
    return wb

# ------------------------------------------------------------------------------
# STREAMLIT UI & MAIN LOGIC
# ------------------------------------------------------------------------------
st.markdown("### 📍 City & File Selection")

c_col1, c_col2 = st.columns([1, 2])

with c_col1:
    city_option = st.selectbox(
        "Select Report City:",
        options=["Kamoke", "Nowshera Virkan", "Gujranwala", "Other (Custom)"]
    )
    if city_option == "Other (Custom)":
        selected_city = st.text_input("Enter City Name:", value="Kamoke").strip().upper()
    else:
        selected_city = city_option.upper()

col_up1, col_up2 = st.columns(2)

with col_up1:
    current_file = st.file_uploader("1️⃣ Upload TODAY'S Mileage Report (.xlsx)", type=["xlsx"])

with col_up2:
    previous_file = st.file_uploader("2️⃣ Upload PREVIOUS Mileage Report (.xlsx) - Optional", type=["xlsx"])

with st.expander("⚙️ Optional: Master List Override"):
    rural_override_file = st.file_uploader("Upload New Master List (.xlsx)", type=["xlsx"])

if current_file:
    try:
        df_current_raw = pd.read_excel(current_file, header=None)
        h_idx, c_reg_idx, c_per_idx = process_mileage_df(df_current_raw)
        
        if h_idx == -1:
            st.error("❌ Could not auto-detect 'Reg#' and 'Period' columns in Today's Mileage Report.")
        else:
            # STRICT NORMALIZED RURAL SET BUILDING
            normalized_rural_set = {normalize_reg(x) for x in DEFAULT_RURAL_LIST if normalize_reg(x)}

            if rural_override_file:
                df_rural_raw = pd.read_excel(rural_override_file)
                normalized_rural_set.clear()
                for col in df_rural_raw.columns:
                    for val in df_rural_raw[col].dropna():
                        val_str = str(val).replace('\xa0', '').strip().upper()
                        if val_str and val_str not in ['TROLLY NO.', 'UC NO.']:
                            normalized_rural_set.add(normalize_reg(val_str))

            # Previous File Checking
            prev_vehicles = set()
            if previous_file:
                df_prev_raw = pd.read_excel(previous_file, header=None)
                p_h_idx, p_reg_idx, p_per_idx = process_mileage_df(df_prev_raw)
                if p_h_idx != -1:
                    df_prev_data = df_prev_raw.iloc[p_h_idx + 1:].copy()
                    for _, p_row in df_prev_data.iterrows():
                        p_hrs = parse_period_to_hours(p_row.iloc[p_per_idx])
                        if 20.0 <= p_hrs <= 24.0:
                            p_reg = str(p_row.iloc[p_reg_idx]).replace('\xa0', '').strip()
                            if p_reg and p_reg not in ['-', 'NAN', 'NONE']:
                                prev_vehicles.add(normalize_reg(p_reg))

            meta_date = extract_file_date(df_current_raw)

            raw_headers = df_current_raw.iloc[h_idx].astype(str).str.replace('\xa0', '', regex=False).str.strip().tolist()
            has_sno = raw_headers[0].upper() in ['S.NO', 'SR.#', 'SR NO', 'S#', 'NO']
            headers = raw_headers if has_sno else ['S.No'] + raw_headers

            df_data = df_current_raw.iloc[h_idx + 1:].copy()
            
            urban_rows, rural_rows = [], []
            reg_counts = {}
            missing_reg_rows = []
            repeated_vehicles = set()
            
            for idx, row in df_data.iterrows():
                period_val = row.iloc[c_per_idx]
                hrs = parse_period_to_hours(period_val)
                
                if 20.0 <= hrs <= 24.0:
                    reg_val = str(row.iloc[c_reg_idx]).replace('\xa0', '').strip() if pd.notna(row.iloc[c_reg_idx]) else ''
                    
                    if not reg_val or reg_val in ['-', 'NAN', 'NONE', ''] or len(reg_val) < 2:
                        missing_reg_rows.append(row.tolist())
                    else:
                        reg_clean = reg_val.upper()
                        reg_norm = normalize_reg(reg_clean)
                        reg_counts[reg_clean] = reg_counts.get(reg_clean, 0) + 1
                        
                        if reg_norm in prev_vehicles:
                            repeated_vehicles.add(reg_clean)
                        
                        row_list = row.tolist()
                        if not has_sno:
                            row_list = [''] + row_list
                        
                        # STRICT MATCHING CHECK
                        if is_rural_vehicle(reg_clean, normalized_rural_set):
                            rural_rows.append(row_list)
                        else:
                            urban_rows.append(row_list)
            
            duplicate_regs = {k for k, v in reg_counts.items() if v > 1}
            total_valid_vehicles = len(urban_rows) + len(rural_rows)

            # Summary Metrics
            st.markdown("### 📊 Audit Summary")
            st.info(f"📍 Selected Location: **{selected_city}** | 📅 File Data Date: **{meta_date}**")
            
            repeated_count = len(repeated_vehicles)
            new_count = total_valid_vehicles - repeated_count - len(duplicate_regs)
            
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Total Filtered (20–24h)", total_valid_vehicles)
            m2.metric("Urban Fleet", len(urban_rows))
            m3.metric("Rural Fleet", len(rural_rows))
            m4.metric("🔄 Repeated", repeated_count)
            m5.metric("✅ New", new_count if new_count > 0 else 0)

            if repeated_vehicles:
                st.info(f"🔁 **Repeated Vehicles ({len(repeated_vehicles)}):** {', '.join(repeated_vehicles)}")
            if duplicate_regs:
                st.warning(f"⚠️ **Duplicate Vehicles ({len(duplicate_regs)}):** {', '.join(duplicate_regs)}")

            # Generate Report
            wb = generate_professional_excel(
                headers, urban_rows, rural_rows, selected_city, 
                meta_date, total_valid_vehicles, repeated_vehicles, 
                duplicate_regs, has_sno, c_reg_idx, prev_vehicles
            )
            
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            
            st.success("🎉 Executive Audit Report Generated Successfully!")
            st.download_button(
                label=f"📥 Download {selected_city} Executive Report (.xlsx)",
                data=output_buffer,
                file_name=f"{selected_city}_Mileage_Report_{meta_date.replace('/', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"Error processing report: {str(e)}")
